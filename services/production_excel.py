from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("generated_reports")


HEADER_FILL = "D9EAF7"
TITLE_FILL = "1F4E78"
ORDER_HEADERS = [
    "OA ID",
    "OA Date",
    "Payment Terms",
    "Purchase Order Date",
    "PO Number",
    "Client Name",
    "State",
    "Item Code",
    "Quantity",
    "Rate",
    "Discount (%)",
    "Amount (₹)",
]


BILLED_HEADERS = [
    "OA ID",
    "OA Date",
    "Payment Terms",
    "Purchase Order Date",
    "PO Number",
    "Client Name",
    "State",
    "Item Code",
    "Quantity",
    "Rate",
    "Discount (%)",
    "Amount (₹)",
    "Bill Number",
    "Bill Date",
    "Billed Quantity",
]

def style_worksheet(
    worksheet,
    title,
    period,
    headers,
    rows,
    widths,
):
    # =========================================================
    # TITLE
    # =========================================================

    worksheet["A1"] = title

    worksheet["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF",
    )

    worksheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor=TITLE_FILL,
    )

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(headers),
    )

    worksheet["A2"] = (
        f"Period: {period[0]} to {period[1]}"
    )

    worksheet["A2"].font = Font(
        italic=True,
        color="666666",
    )

    # =========================================================
    # HEADERS
    # =========================================================

    header_row = 4

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.font = Font(
            bold=True,
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_FILL,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # =========================================================
    # DATA
    # =========================================================

    for row_index, row in enumerate(
        rows,
        start=header_row + 1,
    ):
        for column_index, value in enumerate(
            row,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="center",
            )

    # =========================================================
    # FORMATTING
    # =========================================================

    worksheet.freeze_panes = "A5"

    if rows:
        worksheet.auto_filter.ref = (
            f"A{header_row}:"
            f"{get_column_letter(len(headers))}"
            f"{worksheet.max_row}"
        )

    for column_index, width in widths.items():
        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[4].height = 30

# =============================================================
# Workbook generator
# =============================================================

def order_row(item):
    return [
        item.get("oa_id"),
        item.get("oa_date"),
        item.get("payment_terms"),
        item.get("purchase_order_date"),
        item.get("po_number"),
        item.get("client_name"),
        item.get("state_name"),
        item.get("item_code"),
        item.get("quantity"),
        item.get("rate"),
        item.get("discount_percentage", 0),
        item.get("amount", 0),
    ]

def billed_row(item):
    return [
        item.get("oa_id"),
        item.get("oa_date"),
        item.get("payment_terms"),
        item.get("purchase_order_date"),
        item.get("po_number"),
        item.get("client_name"),
        item.get("state_name"),
        item.get("item_code"),
        item.get("quantity"),
        item.get("rate"),
        item.get("discount_percentage", 0),
        item.get("amount", 0),
        item.get("bill_num"),
        item.get("bill_date"),
        item.get("billed_quantity"),
    ]

def generate_production_excel(
    analytics_data,
    from_date,
    to_date,
):
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    period = (
        from_date,
        to_date,
    )

    common_widths = {
        1: 20,  # OA ID
    2: 16,  # OA Date
    3: 22,  # Payment Terms
    4: 20,  # Purchase Order Date
    5: 20,  # PO Number
    6: 30,  # Client Name
    7: 18,  # State
    8: 20,  # Item Code
    9: 14,  # Quantity
    10: 16, # Rate
    11: 15, # Discount
    12: 18, # Amount
    }

    billed_widths = {
        **common_widths,
    13: 18, # Bill Number
    14: 16, # Bill Date
    15: 18, # Billed Quantity
    }

    # =========================================================
    # 1. ORDERS BOOKED
    # =========================================================

    worksheet = workbook.create_sheet(
        "Orders Booked"
    )

    orders_booked = analytics_data.get(
        "orders_booked",
        [],
    )

    style_worksheet(
        worksheet=worksheet,
        title="Orders Booked",
        period=period,
        headers=ORDER_HEADERS,
        rows=[
            order_row(item)
            for item in orders_booked
        ],
        widths=common_widths,
    )

    # =========================================================
    # 2. PENDING ORDERS
    # =========================================================

    worksheet = workbook.create_sheet(
        "Pending Orders"
    )

    pending_orders = analytics_data.get(
        "pending_order_items",
        [],
    )

    style_worksheet(
        worksheet=worksheet,
        title="Pending Orders",
        period=period,
        headers=ORDER_HEADERS,
        rows=[
            order_row(item)
            for item in pending_orders
        ],
        widths=common_widths,
    )

    # =========================================================
    # 3. BILLED
    # =========================================================

    worksheet = workbook.create_sheet(
        "Billed"
    )

    billed_items = analytics_data.get(
        "billed_items",
        [],
    )

    style_worksheet(
        worksheet=worksheet,
        title="Billed",
        period=period,
        headers=BILLED_HEADERS,
        rows=[
            billed_row(item)
            for item in billed_items
        ],
        widths=billed_widths,
    )

    # =========================================================
    # 4. ORDERED & BILLED
    # =========================================================

    worksheet = workbook.create_sheet(
        "Ordered & Billed"
    )

    ordered_and_billed = analytics_data.get(
        "ordered_and_billed",
        [],
    )

    style_worksheet(
        worksheet=worksheet,
        title="Ordered & Billed",
        period=period,
        headers=BILLED_HEADERS,
        rows=[
            billed_row(item)
            for item in ordered_and_billed
        ],
        widths=billed_widths,
    )

    # =========================================================
    # NUMBER FORMATTING
    # =========================================================

    for worksheet in workbook.worksheets:

        for row in range(5, worksheet.max_row + 1):

            # Rate
            worksheet.cell(
                row=row,
                column=10,
            ).number_format = '₹#,##0.00'

            # Discount
            worksheet.cell(
                row=row,
                column=11,
            ).number_format = '0.00'

            # Amount
            worksheet.cell(
                row=row,
                column=12,
            ).number_format = '₹#,##0.00'
            
    # =========================================================
    # SAVE
    # =========================================================

    filename = (
        f"Production_Analytics_"
        f"{from_date}_to_{to_date}.xlsx"
    )

    output_path = OUTPUT_DIR / filename

    workbook.save(output_path)

    return output_path