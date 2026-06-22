from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
import base64
from services.ai_vision_parser import extract_grn_from_image, calculate_grn_totals, process_grn_image
from security import verify_bearer_token
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
import io
from io import BytesIO
from database.repository import EDBR
from openpyxl.styles import (Font, PatternFill, Border, Side, Alignment)
import csv
import datetime

router = APIRouter(prefix="/api/v1/wms", tags=["Warehouse & Inventory"])

def normalize_item(item: dict) -> dict:
    qty = float(item.get("quantity") or 0)
    rate = float(item.get("rate") or 0)
    disc_pct = float(item.get("discount_percent") or 0)

    gross = qty * rate
    discount_amount = gross * disc_pct / 100
    net = gross - discount_amount

    return {
        "item_code": (item.get("item_code") or "").strip(),

        # Vendor bill values
        "item_name": item.get("item_name", ""),
        "description": item.get("description", ""),

        # ERP matched description
        "item_description": item.get("item_description", ""),

        "quantity": qty,
        "rate": rate,
        "discount_percent": disc_pct,

        "gross_amount": round(gross, 2),
        "discount_amount": round(discount_amount, 2),
        "net_amount": round(net, 2)
    }

def hydrate_items_with_master(items: list) -> list:

    hydrated = []

    for item in items:

        code = (item.get("item_code") or "").strip()

        item["item_description"] = ""

        if not code:
            hydrated.append(item)
            continue

        master = None

        try:
            master = EDBR.get_test_item_by_code(code)
        except Exception:
            master = None

        if master:
            item["item_description"] = master.get("item_specification") or ""

        hydrated.append(item)

    return hydrated

@router.post("/grn/scan-bill")
async def scan_vendor_bill(file: UploadFile = File(...), user: dict = Depends(verify_bearer_token)):
    # Restrict to images for the vision model
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Please upload an image file (JPG/PNG) of the bill.")

    try:
        # Read file and convert to Base64
        file_bytes = await file.read()
        base64_encoded = base64.b64encode(file_bytes).decode('utf-8')
        
        # Pass to Groq Vision
        ai_extracted_data = process_grn_image(base64_encoded)
        subtotal = 0

        for item in ai_extracted_data["items"]:
            qty = float(item.get("quantity", 0) or 0)
            rate = float(item.get("rate", 0) or 0)

            item["amount"] = round(qty * rate, 2)
            subtotal += item["amount"]

            cgst = round(subtotal * 0.09, 2)
            sgst = round(subtotal * 0.09, 2)

            ai_extracted_data["taxes"] = {
                "cgst": cgst,
                "sgst": sgst
            }
        ai_extracted_data["subtotal"] = round(subtotal, 2)
        ai_extracted_data["grand_total"] = round(subtotal + cgst + sgst, 2)

        return {
            "status": "success",
            "message": "Bill scanned successfully.",
            "data": ai_extracted_data
        }
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"Failed to process scan: {str(e)}")

@router.post("/grn/save")
async def save_grn(payload: dict, user: dict = Depends(verify_bearer_token)):
    result = EDBR.create_grn(payload, user["email"])

    return {
        "success": True,
        **result
    }

@router.get("/grn/export/{grn_id}")
async def scan_vendor_bill(file: UploadFile = File(...), user: dict = Depends(verify_bearer_token)):

    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Upload image only")

    file_bytes = await file.read()
    base64_encoded = base64.b64encode(file_bytes).decode("utf-8")

    data = extract_grn_from_image(base64_encoded)

    items = [normalize_item(i) for i in data.get("items", [])]

    subtotal = sum(i["net_amount"] for i in items)

    cgst = round(subtotal * 0.09, 2)
    sgst = round(subtotal * 0.09, 2)

    data["items"] = items
    data["subtotal"] = round(subtotal, 2)
    data["taxes"] = {"cgst": cgst, "sgst": sgst}
    data["grand_total"] = round(subtotal + cgst + sgst, 2)

    return {
        "status": "success",
        "message": "Bill scanned successfully",
        "data": data
    }


# ==============================
# SAVE GRN
# ==============================
@router.post("/grn/save")
async def save_grn(payload: dict, user: dict = Depends(verify_bearer_token)):
    result = EDBR.create_grn(payload, user["email"])
    return {"success": True, **result}


# ==============================
# EXPORT GRN
# ==============================
@router.get("/grn/export/{grn_id}")
async def export_grn_excel(grn_id: int, user: dict = Depends(verify_bearer_token)):

    grn = EDBR.get_grn_by_id(grn_id)

    if not grn:
        raise HTTPException(status_code=404, detail="GRN not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "GRN"

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center = Alignment(horizontal="center", vertical="center")

    headers = ["Sr", "Code", "Desc", "Qty", "Rate", "Gross", "Disc%", "Disc Amt", "Net"]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.border = thin
        c.alignment = center

    row = 2
    subtotal = 0

    for idx, item in enumerate(grn["items"], 1):

        qty = float(item.get("quantity", 0))
        rate = float(item.get("rate", 0))
        disc = float(item.get("discount_percent", 0))

        gross = qty * rate
        disc_amt = gross * disc / 100
        net = item.get("net_amount", gross - disc_amt)

        subtotal += net

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=item.get("item_code", ""))
        ws.cell(row=row, column=3, value=item.get("description", ""))
        ws.cell(row=row, column=4, value=qty)
        ws.cell(row=row, column=5, value=rate)
        ws.cell(row=row, column=6, value=gross)
        ws.cell(row=row, column=7, value=disc)
        ws.cell(row=row, column=8, value=disc_amt)
        ws.cell(row=row, column=9, value=net)

        row += 1

    cgst = round(subtotal * 0.09, 2)
    sgst = round(subtotal * 0.09, 2)

    row += 2
    ws.cell(row=row, column=8, value="Subtotal")
    ws.cell(row=row, column=9, value=subtotal)

    ws.cell(row=row + 1, column=8, value="CGST")
    ws.cell(row=row + 1, column=9, value=cgst)

    ws.cell(row=row + 2, column=8, value="SGST")
    ws.cell(row=row + 2, column=9, value=sgst)

    ws.cell(row=row + 3, column=8, value="Grand Total")
    ws.cell(row=row + 3, column=9, value=subtotal + cgst + sgst)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={grn['grn_number']}.xlsx"}
    )

@router.post("/grn/export-preview")
async def export_grn_preview(payload: dict, user: dict = Depends(verify_bearer_token)):

    wb = Workbook()
    ws = wb.active
    ws.title = "GRN PREVIEW"

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:D1")
    ws["A1"] = "Tempo Instruments Pvt. Ltd."
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = center

    ws["E1"] = "GRN No:"
    ws["F1"] = payload.get("grn_number", "")

    ws["E2"] = "Date:"
    ws["F2"] = payload.get("invoice_date", "")

    ws["A3"] = "Vendor:"
    ws.merge_cells("B3:F3")
    ws["B3"] = payload.get("vendor_name", "")

    start_row = 6

    headers = ["Sr", "Code", "Description", "Qty", "Rate", "Gross", "Disc %", "Disc Amt", "Net"]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = Font(bold=True)
        c.border = thin
        c.alignment = center

    items = [normalize_item(i) for i in payload.get("items", [])]

    row = start_row + 1

    
    for idx, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=item["item_code"])
        display_text = item.get("item_name", "")
        if item.get("item_description"):
            display_text += f"\n{item['item_description']}"
        cell = ws.cell(row=row, column=3, value=display_text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=4, value=item["quantity"])
        ws.cell(row=row, column=5, value=item["rate"])
        ws.cell(row=row, column=6, value=item["gross_amount"])
        ws.cell(row=row, column=7, value=item["discount_percent"])
        ws.cell(row=row, column=8, value=item["discount_amount"])
        ws.cell(row=row, column=9, value=item["net_amount"])
        row += 1

    shipping = float(payload.get("shipping", 0))
    totals = calculate_grn_totals(items, shipping)

    row += 2
    ws.cell(row=row, column=8, value="Shipping charges:")
    ws.cell(row=row, column=9, value=shipping)

    ws.cell(row=row+1, column=8, value="Subtotal")
    ws.cell(row=row+1, column=9, value=totals["subtotal"])

    ws.cell(row=row+2, column=8, value="CGST")
    ws.cell(row=row+2, column=9, value=totals["taxes"]["cgst"])

    ws.cell(row=row+3, column=8, value="SGST")
    ws.cell(row=row+3, column=9, value=totals["taxes"]["sgst"])

    ws.cell(row=row+4, column=8, value="Grand Total")
    ws.cell(row=row+4, column=9, value=totals["grand_total"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=GRN_PREVIEW.xlsx"}
    )

@router.post("/items/seed-test-csv")
async def seed_test_items_from_csv(file: UploadFile = File(...), user: dict = Depends(verify_bearer_token)):
    # 1. Validate file type
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are supported for this operation.")

    try:
        # 2. Read and decode the CSV file in memory
        contents = await file.read()
        
        try:
            decoded = contents.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded = contents.decode('cp1252')

        reader = csv.DictReader(io.StringIO(decoded))
        print("Headers:", reader.fieldnames)
        items_to_insert = []
        
        # 3. Map the CSV rows to our database schema
        for row in reader:
            # We use .get() to safely pull the headers you mentioned
            item_code = row.get('Item code', '').strip()
            item_spec = row.get('Item Specifications', '').strip()

            if item_code:
                items_to_insert.append({
                    "item_code": item_code,
                    "item_specification": item_spec
                })

        if not items_to_insert:
             raise ValueError("No valid items found. Please ensure headers are exactly 'Item code' and 'Item Specifications'.")

        # 4. Trigger the bulk database insert
        EDBR.seed_test_items(items_to_insert)

        return {
            "status": "success", 
            "message": f"Successfully processed the CSV. Sent {len(items_to_insert)} items to the testing database."
        }

    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"Failed to process CSV upload: {str(e)}")
    
@router.get("/test-item/{item_code}")
def get_test_item(item_code: str):
    item = EDBR.get_test_item_by_code(item_code)

    if not item:
        return None

    return item